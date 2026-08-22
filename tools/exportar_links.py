#!/usr/bin/env python3
# ═══════════════════════════════════════════════════════════════════
#  EXPORTAR LINKS → hoja de cálculo (CSV + XLSX, 2 hojas Mac/Windows)
#
#  Corre en el proyecto ORIGINAL (donde vive el catálogo con las URLs).
#  Genera:
#    - links_para_sheet.csv   → para pegar directo en la hoja "Links"
#    - links_catalogo.xlsx    → Excel con 2 hojas ("Links Mac"/"Links Windows"),
#                               agrupado por tipo (Apps / Adobe / Herramientas)
#
#  Columnas: nombre | metodo | plataforma | url | resolver | tipo | para_apps
#    - tipo:      Apps / Adobe / Herramientas (visual, no la usa el script)
#    - para_apps: apps que usan cada herramienta (p. ej. "Steinberg Patcher")
#  Las 5 primeras son las del Apps Script; las extras no estorban.
#
#  Cubre TODO lo que el wizard pide por get_link en modo servidor:
#    - Apps regulares  (catalog/urls.py, HTTP directo + magnets + fallback)
#    - Apps Adobe      (por método: aio_macked / aio_sice / multilang_sice;
#                       usa el link ARM; editá la hoja para Intel si hace falta)
#    - Tools de método (ADOBE_TOOLS por método: Sentinel, AntiCC, ...)
#    - Tools por app   (patchers Steinberg/Adobe, core de Office, etc.)
#
#  Uso:  python3 tools/exportar_links.py
#
#  IMPORTANTE AL USARLO EN GOOGLE SHEETS: el Apps Script lee UNA hoja
#  llamada "Links". Al agregar el .xlsx, pegá AMBAS hojas (o su contenido
#  concatenado) en esa "Links": la columna "plataforma" distingue mac/win.
# ═══════════════════════════════════════════════════════════════════

import csv
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from catalog.urls import (  # noqa: E402
    _DOWNLOAD_URLS_MAC, _DOWNLOAD_URLS_WIN,
    _TORRENT_MAGNETS_MAC, _TORRENT_MAGNETS_WIN,
    SWISSTRANSFER_URLS,
)
from catalog.adobe import (  # noqa: E402
    ADOBE_AIO_MACKED_LINKS, ADOBE_AIO_SICE_LINKS, ADOBE_MULTILANG_LINKS,
    ADOBE_TOOLS, ADOBE_APPS,
)
from catalog.adobe_helpers import _adobe_link_flat  # noqa: E402
from catalog.specs import DOWNLOAD_METHODS  # noqa: E402
from catalog.tools import _app_tools_for_app  # noqa: E402

HEADERS = ["nombre", "metodo", "plataforma", "url", "resolver", "tipo", "para_apps"]

TIPO_APPS = "Apps"
TIPO_ADOBE = "Adobe"
TIPO_HERRAMIENTAS = "Herramientas"
TIPO_ORDER = {TIPO_APPS: 0, TIPO_ADOBE: 1, TIPO_HERRAMIENTAS: 2}

ADOBE_SOURCES = {
    "aio_macked": ADOBE_AIO_MACKED_LINKS,
    "aio_sice": ADOBE_AIO_SICE_LINKS,
    "multilang_sice": ADOBE_MULTILANG_LINKS,
}


def resolver_for(url: str) -> str:
    """Infiere el resolver según el host de la URL."""
    u = (url or "").lower()
    if not u:
        return ""
    if "pixeldrain.com" in u:
        return "pixeldrain"
    if "akirabox" in u:
        return "akirabox"
    if "workupload" in u:
        return "workupload"
    if "seyarabata" in u:
        return "seyarabata"
    if "appstorrent" in u:
        return "appstorrent"
    if "swisstransfer" in u:
        return "swisstransfer"
    return ""


def _merged_apps(current: str, new) -> str:
    """Une apps (columna para_apps) sin repetir: 'A, B' + 'B, C' → 'A, B, C'."""
    items = [p.strip() for p in (current or "").split(",") if p and p.strip()]
    for n in (new if isinstance(new, list) else [new]):
        n = (n or "").strip()
        if n and n not in items:
            items.append(n)
    return ", ".join(items)


def build_rows():
    rows = []
    seen = set()
    by_key = {}

    def add(name, platform, url, method="", tipo=TIPO_APPS, para=""):
        url = (url or "").strip()
        if not url or url == "combo":
            return
        para = ", ".join(para) if isinstance(para, list) else (para or "")
        key = (name, method, platform)
        if key in seen:
            if para and key in by_key:
                by_key[key][6] = _merged_apps(by_key[key][6], para)
            return  # primera fila gana (URL), pero se suman las apps
        seen.add(key)
        row = [name, method, platform, url, resolver_for(url), tipo, para]
        rows.append(row)
        by_key[key] = row

    # 1) Apps regulares (HTTP directo) + magnets + fallback SwissTransfer.
    for name, url in _DOWNLOAD_URLS_MAC.items():
        add(name, "mac", url)
    for name, url in _DOWNLOAD_URLS_WIN.items():
        add(name, "win", url)
    for name, url in _TORRENT_MAGNETS_MAC.items():
        add(name, "mac", url)
    for name, url in _TORRENT_MAGNETS_WIN.items():
        add(name, "win", url)
    for name, url in SWISSTRANSFER_URLS.items():
        for platform in ("mac", "win"):
            if (name, platform) not in by_key:
                add(name, platform, url)

    # 2) Apps Adobe por método (link ARM de la versión más nueva).
    for method, source in ADOBE_SOURCES.items():
        for app, entry in source.items():
            versions = entry if isinstance(entry, list) else [entry]
            if not versions:
                continue
            url = _adobe_link_flat(versions[0], "arm")
            add(app, "mac", url, method=method, tipo=TIPO_ADOBE)

    # 3) Tools de método (ADOBE_TOOLS × for_methods), para mac y win.
    for tool_name, cfg in ADOBE_TOOLS.items():
        for method in cfg.get("for_methods", []):
            for platform in ("mac", "win"):
                add(tool_name, platform, cfg.get("url", ""), method=method,
                    tipo=TIPO_HERRAMIENTAS, para=["Adobe (todas)"])

    # 4) Tools por app (patchers Steinberg, core de Office, etc.), método vacío.
    #    Se omiten los patchers por programa Adobe (source=adobe_patcher): el
    #    wizard ya no los pide (las apps Adobe van por método o por torrent).
    for app in set(DOWNLOAD_METHODS) | set(ADOBE_APPS):
        for tool in _app_tools_for_app(app):
            if tool.get("source") == "adobe_patcher":
                continue
            name = tool.get("name", app)
            url = tool.get("url", "")
            for platform in ("mac", "win"):
                add(name, platform, url, method="", tipo=TIPO_HERRAMIENTAS,
                    para=[app])

    return rows


def sort_rows(rows):
    """Agrupa por plataforma → tipo (Apps, Adobe, Herramientas) → orden."""
    def key(r):
        return (
            r[2],                       # plataforma
            TIPO_ORDER.get(r[5], 99),   # tipo: primero Apps, luego Adobe,
                                        # luego Herramientas
            r[1] if r[5] == TIPO_ADOBE else "",  # Adobe agrupado por método
            r[0],                       # nombre
        )
    return sorted(rows, key=key)


def write_csv(rows):
    out = ROOT / "links_para_sheet.csv"
    with open(out, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(r for r in rows)
    return out


def write_xlsx(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="2A2A2A")
    header_font = Font(bold=True, color="FFFFFF")
    link_font = Font(color="0563C1", underline="single")
    tipo_fill = {
        TIPO_APPS: PatternFill("solid", fgColor="D9EAD3"),          # verde
        TIPO_ADOBE: PatternFill("solid", fgColor="FCE5CD"),         # naranja
        TIPO_HERRAMIENTAS: PatternFill("solid", fgColor="CFE2F3"),  # azul
    }

    for platform in ("mac", "win"):
        ws = wb.create_sheet(f"Links {platform.capitalize()}")
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in (r for r in rows if r[2] == platform):
            ws.append(row)
            ws.cell(row=ws.max_row, column=6).fill = tipo_fill.get(row[5])
        for i, width in enumerate([30, 16, 12, 72, 14, 14, 46], start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        for row in ws.iter_rows(min_row=2):
            url_cell = row[3]
            url_cell.hyperlink = url_cell.value
            url_cell.style = "Hyperlink"
            url_cell.font = link_font
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:G{ws.max_row}"

    out = ROOT / "links_catalogo.xlsx"
    wb.save(out)
    return out


def main():
    rows = sort_rows(build_rows())
    csv_out = write_csv(rows)
    xlsx_out = write_xlsx(rows)

    print(f"Filas totales: {len(rows)}")
    print(f"  CSV  → {csv_out}")
    print(f"  XLSX → {xlsx_out}  (hojas: Links Mac / Links Windows)")
    by = {"mac": 0, "win": 0}
    for r in rows:
        by[r[2]] = by.get(r[2], 0) + 1
    print(f"    mac: {by['mac']} filas | win: {by['win']} filas")
    print("Métodos:")
    met = {}
    for r in rows:
        met[r[1] or "(vacío)"] = met.get(r[1] or "(vacío)", 0) + 1
    for k, v in sorted(met.items()):
        print(f"  {k}: {v}")
    print("Resolver por fila:")
    res = {}
    for r in rows:
        res[r[4]] = res.get(r[4], 0) + 1
    for k, v in sorted(res.items()):
        print(f"  {k or '(directo)'}: {v}")


if __name__ == "__main__":
    main()
